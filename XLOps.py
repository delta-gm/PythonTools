import openpyxl
from openpyxl import Workbook


class XLOps:

    def __init__(self):

        # alphabet is used to convert a number 1-650 to a column by letter
        self.alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        self.alphabet = ['']
        last = 650
        for k in range(0, last):
            if k < 26:
                self.alphabet.append(self.alpha[k])
            else:
                num = int(k / 26)
                rem = k % 26
                self.alphabet.append(self.alpha[num-1] + self.alpha[rem])

    # Format the columns of a worksheet to fit to width
    def fit_to_width(self,
                     ws,  # Openpyxl worksheet
                     ):
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = max(value + 4, 6)

    # Write a list of lists to a worksheet
    def write_sheet(self,
                    ws,  # Openpyxl worksheet
                    in_list,  # Data as a matrix (list of lists)
                    corner,  # Top left corner in format ['A', 1]
                    ):
        startcol = self.alphabet.index(corner[0])
        startrow = corner[1]
        for r in range(0, len(in_list)):
            for c in range(0, len(in_list[0])):
                ws.cell(row=(startrow + r), column=(startcol + c)).value = in_list[r][c]

    # Read data below a header that is an EXACT MATCH for a key string (not case sensitive)
    # read_sheet returns a list of lists (row, column), but read_column_by_header returns a column as a single list
    def read_column_by_header(self,
                              ws,  # Openpyxl worksheet
                              key,  # Desired header
                              ):
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                if ws.cell(row=r, column=c).value is not None:
                    if key.upper() == str(ws.cell(row=r, column=c).value).upper():
                        return [k[0] for k in self.read_sheet(ws, [self.alphabet[c], r, self.alphabet[c], ws.max_row])]

    # Read a worksheet into a list of lists
    def read_sheet(self,
                   ws,  # Openpyxl worksheet
                   xrange,  # Data range in format ['A', 1, 'Z', 100]
                   ):
        startcol = self.alphabet.index(xrange[0])
        startrow = xrange[1]
        endcol = self.alphabet.index(xrange[2])
        endrow = xrange[3]
        out_list = [None] * (endrow - startrow + 1)
        for k in range(0, len(out_list)):
            out_list[k] = [None] * (endcol - startcol + 1)
        for r in range(0, endrow - startrow + 1):
            for c in range(0, endcol - startcol + 1):
                out_list[r][c] = str(ws.cell(row=(startrow + r), column=(startcol + c)).value)
                if out_list[r][c] == 'None':
                    out_list[r][c] = ''
        return out_list

    # Returns a list of cell indices representing the top left corner position of data sets to be written in blocks.
    # Useful for building a calendar or nested tables in Excel.
    def get_block_points(self,
                         start,  # Starting cell in format ['A', 1]
                         block_width,  # Width of inner block, in cells
                         block_height,  # Height of inner block, in cells
                         fig_width,  # Width of outer figure, in blocks
                         fig_height,  # Height of outer figure, in blocks
                         ):
        out_list = []
        startcol = self.alphabet.index(start[0])
        startrow = start[1]
        for r in range(0, fig_height):
            for c in range(0, fig_width):
                out_list.append([self.alphabet[startcol + c*block_width], startrow + r*block_height])
        return out_list  # Upper left corners in format [['A', 1], ['B', 2]]

'''

# Sample code

A = XLOps()
wb = Workbook()
ws = wb.worksheets[0]
mydata = [[  'Fruit', 'Qty'],
          [ 'Apples',     6],
          ['Bananas',     3],
          ['Oranges',    12],
          ]
A.write_sheet(ws, mydata, ['A', 1])
newdata = A.read_sheet(ws, ['A', 1, 'B', 4])
for k in newdata:
    print(k)
fruits = A.read_column_by_header(ws, 'Fruit')[1:]
print(fruits)
wb.save('Sample Worksheet.xlsx')
'''
